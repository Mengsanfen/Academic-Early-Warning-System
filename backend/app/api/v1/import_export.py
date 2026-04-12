"""
学生批量导入 API
"""
import io
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
import openpyxl
from openpyxl.workbook import Workbook

from app.database import get_db
from app.api.deps import check_class_access, get_admin_user, get_counselor_or_admin
from app.models.user import User, UserRole
from app.models.student import Student, Class, StudentStatus, Department
from app.models.score import Score
from app.models.attendance import Attendance, AttendanceStatus
from app.models.course import Course
from app.core.security import get_password_hash


router = APIRouter()


def _ensure_import_scope(current_user: User, student: Student, course: Course) -> None:
    """Ensure counselors can only import data for managed classes."""
    student_class_id = student.class_id

    if student_class_id and course.class_id is None:
        course.class_id = student_class_id

    if current_user.role != UserRole.COUNSELOR:
        return

    if not student_class_id or not check_class_access(current_user, student_class_id):
        raise PermissionError("无权导入该学生所在班级的数据")

    if not course.class_id:
        raise PermissionError("课程未绑定班级，无法确认导入权限")

    if course.class_id != student_class_id:
        raise PermissionError("课程所属班级与学生所在班级不一致，无法导入")

    if not check_class_access(current_user, course.class_id):
        raise PermissionError("无权导入该课程所属班级的数据")


def get_or_create_department(db: Session, name: str = "默认院系") -> Department:
    """获取或创建院系"""
    dept = db.query(Department).filter(Department.name == name).first()
    if not dept:
        dept = Department(name=name, code="DEFAULT")
        db.add(dept)
        db.flush()
    return dept


# ========== Schemas ==========

class ImportResult(BaseModel):
    """导入结果"""
    total: int
    success: int
    failed: int
    errors: List[str] = []
    created_users: int = 0


class StudentImportRow(BaseModel):
    """学生导入行数据"""
    student_no: str
    name: str
    gender: Optional[str] = None
    class_name: str
    enroll_year: Optional[int] = None
    phone: Optional[str] = None
    email: Optional[str] = None


# ========== API 端点 ==========

@router.post("/import", response_model=ImportResult, summary="批量导入学生")
async def import_students(
    file: UploadFile = File(...),
    create_accounts: bool = True,
    default_password_type: str = "student_no",  # student_no 或 phone
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    """
    批量导入学生数据

    Excel格式要求（第一行为表头）：
    - 学号 (必填)
    - 姓名 (必填)
    - 性别
    - 班级名称 (必填)
    - 入学年份
    - 手机号
    - 邮箱

    参数:
    - create_accounts: 是否同时创建用户账号
    - default_password_type: 默认密码类型 (student_no=学号, phone=手机号后6位)
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="只支持 Excel 文件 (.xlsx, .xls)"
        )

    # 读取Excel文件
    content = await file.read()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Excel 文件解析失败: {str(e)}"
        )

    result = ImportResult(total=0, success=0, failed=0, errors=[], created_users=0)

    # 获取表头
    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value).strip() if cell.value else "")

    # 检查必填列
    required_columns = ["学号", "姓名", "班级名称"]
    for col in required_columns:
        if col not in headers:
            result.errors.append(f"缺少必填列: {col}")

    if result.errors:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="; ".join(result.errors)
        )

    # 获取列索引
    col_map = {h: i for i, h in enumerate(headers)}

    # 处理数据行
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        result.total += 1

        try:
            # 解析行数据
            student_no = str(row[col_map["学号"]] or "").strip()
            name = str(row[col_map["姓名"]] or "").strip()
            gender = str(row[col_map.get("性别", -1)] or "").strip() if "性别" in col_map else None
            class_name = str(row[col_map["班级名称"]] or "").strip()
            enroll_year = int(row[col_map.get("入学年份", -1)]) if "入学年份" in col_map and row[col_map.get("入学年份", -1)] else None
            phone = str(row[col_map.get("手机号", -1)] or "").strip() if "手机号" in col_map else None
            email = str(row[col_map.get("邮箱", -1)] or "").strip() if "邮箱" in col_map else None

            # 验证必填字段
            if not student_no or not name or not class_name:
                result.failed += 1
                result.errors.append(f"第{row_idx}行: 学号、姓名、班级名称不能为空")
                continue

            # 检查学号是否已存在
            if db.query(Student).filter(Student.student_no == student_no).first():
                result.failed += 1
                result.errors.append(f"第{row_idx}行: 学号 {student_no} 已存在")
                continue

            # 查找或创建班级
            class_obj = db.query(Class).filter(Class.name == class_name).first()
            if not class_obj:
                # 获取或创建默认院系
                default_dept = get_or_create_department(db)
                # 自动创建班级
                class_obj = Class(
                    name=class_name,
                    grade=str(enroll_year or 2024),
                    department_id=default_dept.id
                )
                db.add(class_obj)
                db.flush()

            # 创建学生
            student = Student(
                student_no=student_no,
                name=name,
                gender=gender,
                class_id=class_obj.id,
                enroll_year=enroll_year,
                phone=phone,
                email=email,
                status=StudentStatus.ACTIVE
            )
            db.add(student)
            db.flush()

            # 创建用户账号
            if create_accounts:
                # 检查用户名是否已存在
                if db.query(User).filter(User.username == student_no).first():
                    result.errors.append(f"第{row_idx}行: 用户名 {student_no} 已存在，跳过创建账号")
                else:
                    # 设置默认密码
                    if default_password_type == "phone" and phone and len(phone) >= 6:
                        default_password = phone[-6:]
                    else:
                        default_password = student_no[-6:] if len(student_no) >= 6 else student_no

                    user = User(
                        username=student_no,
                        password_hash=get_password_hash(default_password),
                        role=UserRole.STUDENT,
                        student_id=student.id,
                        is_active=True,
                        first_login=True,
                        email=email
                    )
                    db.add(user)
                    result.created_users += 1

            result.success += 1

        except Exception as e:
            result.failed += 1
            result.errors.append(f"第{row_idx}行: {str(e)}")

    # 提交事务
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"保存数据失败: {str(e)}"
        )

    return result


@router.get("/template", summary="下载导入模板")
def download_template():
    """下载学生导入Excel模板"""
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "学生导入模板"

    # 设置表头
    headers = ["学号", "姓名", "性别", "班级名称", "入学年份", "手机号", "邮箱"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # 添加示例数据
    sample_data = [
        ["2024001", "张三", "男", "计算机2024级1班", 2024, "13800138001", "zhangsan@example.com"],
        ["2024002", "李四", "女", "计算机2024级1班", 2024, "13800138002", "lisi@example.com"],
        ["2024003", "王五", "男", "计算机2024级2班", 2024, "13800138003", "wangwu@example.com"],
    ]
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 设置列宽
    column_widths = [15, 10, 6, 20, 10, 15, 28]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_import_template.xlsx"}
    )


@router.get("/export", summary="导出学生数据")
def export_students(
    class_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user)
):
    """导出学生数据为Excel"""
    query = db.query(Student)

    if class_id:
        query = query.filter(Student.class_id == class_id)

    students = query.all()

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "学生数据"

    # 设置表头
    headers = ["学号", "姓名", "性别", "班级", "入学年份", "手机号", "邮箱", "状态", "是否有账号"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # 填充数据
    for row_idx, student in enumerate(students, 2):
        ws.cell(row=row_idx, column=1, value=student.student_no)
        ws.cell(row=row_idx, column=2, value=student.name)
        ws.cell(row=row_idx, column=3, value=student.gender)
        ws.cell(row=row_idx, column=4, value=student.class_info.name if student.class_info else "")
        ws.cell(row=row_idx, column=5, value=student.enroll_year)
        ws.cell(row=row_idx, column=6, value=student.phone)
        ws.cell(row=row_idx, column=7, value=student.email)
        ws.cell(row=row_idx, column=8, value=student.status.value)
        ws.cell(row=row_idx, column=9, value="是" if student.user else "否")

    # 设置列宽
    column_widths = [12, 10, 6, 20, 10, 15, 25, 10, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students_export.xlsx"}
    )


# ==================== 成绩导入 ====================

@router.post("/scores/import", response_model=ImportResult, summary="批量导入成绩")
async def import_scores(
    file: UploadFile = File(...),
    update_existing: bool = False,  # 是否更新已存在的成绩
    db: Session = Depends(get_db),
    current_user: User = Depends(get_counselor_or_admin)
):
    """
    批量导入成绩数据

    Excel格式要求（第一行为表头）：
    - 学号 (必填)
    - 课程代码 (必填)
    - 课程名称 (必填)
    - 成绩 (必填)
    - 学分
    - 学期 (必填)
    - 考试类型 (期末/补考/重修)

    参数:
    - update_existing: 是否更新已存在的成绩记录
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="只支持 Excel 文件 (.xlsx, .xls)"
        )

    content = await file.read()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Excel 文件解析失败: {str(e)}"
        )

    result = ImportResult(total=0, success=0, failed=0, errors=[])

    # 获取表头
    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value).strip() if cell.value else "")

    # 检查必填列
    required_columns = ["学号", "课程代码", "成绩", "学期"]
    for col in required_columns:
        if col not in headers:
            result.errors.append(f"缺少必填列: {col}")

    if result.errors:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="; ".join(result.errors)
        )

    col_map = {h: i for i, h in enumerate(headers)}

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        result.total += 1

        try:
            student_no = str(row[col_map["学号"]] or "").strip()
            course_code = str(row[col_map["课程代码"]] or "").strip()
            course_name = str(row[col_map.get("课程名称", -1)] or "").strip() if "课程名称" in col_map else course_code
            score_value = row[col_map["成绩"]]
            credit = float(row[col_map.get("学分", -1)]) if "学分" in col_map and row[col_map.get("学分", -1)] else 1.0
            semester = str(row[col_map["学期"]] or "").strip()
            exam_type = str(row[col_map.get("考试类型", -1)] or "期末").strip() if "考试类型" in col_map else "期末"

            # 验证必填字段
            if not student_no or not course_code or score_value is None or not semester:
                result.failed += 1
                result.errors.append(f"第{row_idx}行: 学号、课程代码、成绩、学期不能为空")
                continue

            # 转换成绩
            try:
                score_float = float(score_value)
            except (ValueError, TypeError):
                result.failed += 1
                result.errors.append(f"第{row_idx}行: 成绩格式错误")
                continue

            # 查找学生
            student = db.query(Student).filter(Student.student_no == student_no).first()
            if not student:
                result.failed += 1
                result.errors.append(f"第{row_idx}行: 学号 {student_no} 不存在")
                continue

            # 查找或创建课程
            course = db.query(Course).filter(Course.course_code == course_code).first()
            if not course:
                course = Course(
                    course_code=course_code,
                    course_name=course_name or course_code,
                    credit=credit,
                    semester=semester,
                    class_id=student.class_id,
                )
                _ensure_import_scope(current_user, student, course)
                db.add(course)
                db.flush()
            else:
                _ensure_import_scope(current_user, student, course)

            # 检查是否已存在成绩
            existing_score = db.query(Score).filter(
                Score.student_id == student.id,
                Score.course_id == course.id,
                Score.semester == semester
            ).first()

            if existing_score:
                if update_existing:
                    existing_score.score = score_float
                    existing_score.exam_type = exam_type
                    result.success += 1
                else:
                    result.failed += 1
                    result.errors.append(f"第{row_idx}行: 学号 {student_no} 的课程 {course_code} 成绩已存在")
                continue

            # 创建成绩记录
            new_score = Score(
                student_id=student.id,
                course_id=course.id,
                score=score_float,
                semester=semester,
                exam_type=exam_type
            )
            db.add(new_score)
            result.success += 1

        except Exception as e:
            result.failed += 1
            result.errors.append(f"第{row_idx}行: {str(e)}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"保存数据失败: {str(e)}"
        )

    return result


@router.get("/scores/template", summary="下载成绩导入模板")
def download_score_template():
    """下载成绩导入Excel模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩导入模板"

    headers = ["学号", "课程代码", "课程名称", "成绩", "学分", "学期", "考试类型"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    sample_data = [
        ["2024001", "CS101", "计算机导论", 85, 3, "2024-2025-1", "期末"],
        ["2024001", "MA101", "高等数学", 72, 4, "2024-2025-1", "期末"],
        ["2024002", "CS101", "计算机导论", 58, 3, "2024-2025-1", "补考"],
    ]
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    column_widths = [15, 12, 20, 8, 6, 12, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=score_import_template.xlsx"}
    )


# ==================== 考勤导入 ====================

@router.post("/attendances/import", response_model=ImportResult, summary="批量导入考勤")
async def import_attendances(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_counselor_or_admin)
):
    """
    批量导入考勤数据

    Excel格式要求（第一行为表头）：
    - 学号 (必填)
    - 课程代码 (必填)
    - 课程名称
    - 日期 (必填，格式：YYYY-MM-DD)
    - 状态 (必填：正常/缺勤/迟到/请假)
    - 备注

    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="只支持 Excel 文件 (.xlsx, .xls)"
        )

    content = await file.read()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Excel 文件解析失败: {str(e)}"
        )

    result = ImportResult(total=0, success=0, failed=0, errors=[])

    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value).strip() if cell.value else "")

    required_columns = ["学号", "课程代码", "日期", "状态"]
    for col in required_columns:
        if col not in headers:
            result.errors.append(f"缺少必填列: {col}")

    if result.errors:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="; ".join(result.errors)
        )

    col_map = {h: i for i, h in enumerate(headers)}

    # 状态映射
    status_map = {
        "正常": AttendanceStatus.PRESENT,
        "出勤": AttendanceStatus.PRESENT,
        "present": AttendanceStatus.PRESENT,
        "缺勤": AttendanceStatus.ABSENT,
        "旷课": AttendanceStatus.ABSENT,
        "absent": AttendanceStatus.ABSENT,
        "迟到": AttendanceStatus.LATE,
        "late": AttendanceStatus.LATE,
        "请假": AttendanceStatus.LEAVE,
        "leave": AttendanceStatus.LEAVE,
    }

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        result.total += 1

        try:
            student_no = str(row[col_map["学号"]] or "").strip()
            course_code = str(row[col_map["课程代码"]] or "").strip()
            course_name = str(row[col_map.get("课程名称", -1)] or "").strip() if "课程名称" in col_map else course_code
            date_value = row[col_map["日期"]]
            status_str = str(row[col_map["状态"]] or "").strip()
            remark = str(row[col_map.get("备注", -1)] or "").strip() if "备注" in col_map else None

            # 验证必填字段
            if not student_no or not course_code or not date_value or not status_str:
                result.failed += 1
                result.errors.append(f"第{row_idx}行: 学号、课程代码、日期、状态不能为空")
                continue

            # 解析日期
            from datetime import datetime
            if isinstance(date_value, datetime):
                attendance_date = date_value.date()
            else:
                try:
                    attendance_date = datetime.strptime(str(date_value), "%Y-%m-%d").date()
                except ValueError:
                    result.failed += 1
                    result.errors.append(f"第{row_idx}行: 日期格式错误，应为 YYYY-MM-DD")
                    continue

            # 解析状态
            attendance_status = status_map.get(status_str)
            if not attendance_status:
                result.failed += 1
                result.errors.append(f"第{row_idx}行: 状态 '{status_str}' 无效，应为：正常/缺勤/迟到/请假")
                continue

            # 查找学生
            student = db.query(Student).filter(Student.student_no == student_no).first()
            if not student:
                result.failed += 1
                result.errors.append(f"第{row_idx}行: 学号 {student_no} 不存在")
                continue

            # 查找或创建课程
            course = db.query(Course).filter(Course.course_code == course_code).first()
            if not course:
                course = Course(
                    course_code=course_code,
                    course_name=course_name or course_code,
                    credit=1.0,
                    semester="unknown",
                    class_id=student.class_id,
                )
                _ensure_import_scope(current_user, student, course)
                db.add(course)
                db.flush()
            else:
                _ensure_import_scope(current_user, student, course)

            # 检查是否已存在考勤记录
            existing = db.query(Attendance).filter(
                Attendance.student_id == student.id,
                Attendance.course_id == course.id,
                Attendance.date == attendance_date
            ).first()

            if existing:
                # 更新状态
                existing.status = attendance_status
                existing.remark = remark
                result.success += 1
                continue

            # 创建考勤记录
            new_attendance = Attendance(
                student_id=student.id,
                course_id=course.id,
                date=attendance_date,
                status=attendance_status,
                remark=remark
            )
            db.add(new_attendance)
            result.success += 1

        except Exception as e:
            result.failed += 1
            result.errors.append(f"第{row_idx}行: {str(e)}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"保存数据失败: {str(e)}"
        )

    return result


@router.get("/attendances/template", summary="下载考勤导入模板")
def download_attendance_template():
    """下载考勤导入Excel模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤导入模板"

    headers = ["学号", "课程代码", "课程名称", "日期", "状态", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    from datetime import date
    today = date.today().strftime("%Y-%m-%d")

    sample_data = [
        ["2024001", "CS101", "计算机导论", today, "正常", ""],
        ["2024002", "CS101", "计算机导论", today, "迟到", "迟到10分钟"],
        ["2024003", "CS101", "计算机导论", today, "缺勤", "旷课"],
        ["2024004", "CS101", "计算机导论", today, "请假", "病假"],
    ]
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    column_widths = [15, 12, 20, 12, 8, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_import_template.xlsx"}
    )
