"""
Alert management APIs with strict role-based data scoping.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, status as http_status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session, joinedload

from app.api.deps import (
    check_student_access,
    get_counselor_or_admin,
    get_current_user,
    get_student_user,
)
from app.database import get_db
from app.models.alert import Alert, AlertRecord, AlertStatus
from app.models.rule import (
    AlertLevel,
    COMPREHENSIVE_RULE_CODE,
    COMPREHENSIVE_RULE_MODE,
    Rule,
)
from app.models.student import Student
from app.models.user import User, UserRole


router = APIRouter()


class HandleRequest(BaseModel):
    action: str = Field(..., min_length=1, max_length=100)
    result: Optional[str] = Field(default=None, max_length=1000)


class AlertFeedbackRequest(BaseModel):
    feedback: str = Field(..., min_length=1, max_length=1000)


def _get_rule_display_type(rule: Optional[Rule]) -> Optional[str]:
    if not rule:
        return None
    if rule.code == COMPREHENSIVE_RULE_CODE:
        return "comprehensive"
    if isinstance(rule.conditions, dict) and rule.conditions.get("mode") == COMPREHENSIVE_RULE_MODE:
        return "comprehensive"
    return rule.type.value if rule.type else None


def _get_level_label(level: AlertLevel) -> str:
    mapping = {
        AlertLevel.WARNING: "警告",
        AlertLevel.SERIOUS: "严重",
        AlertLevel.URGENT: "紧急",
    }
    return mapping.get(level, str(level))


def _get_status_label(status: AlertStatus) -> str:
    mapping = {
        AlertStatus.PENDING: "待处理",
        AlertStatus.PROCESSING: "处理中",
        AlertStatus.RESOLVED: "已解决",
        AlertStatus.IGNORED: "已忽略",
    }
    return mapping.get(status, str(status))


def _scoped_alert_query(db: Session, current_user: User):
    query = (
        db.query(Alert)
        .join(Student, Alert.student_id == Student.id)
        .options(
            joinedload(Alert.student).joinedload(Student.class_info),
            joinedload(Alert.rule),
            joinedload(Alert.records).joinedload(AlertRecord.handler).joinedload(User.student),
        )
    )

    if current_user.role == UserRole.ADMIN:
        return query

    if current_user.role == UserRole.COUNSELOR:
        managed_class_ids = current_user.get_managed_class_ids()
        if not managed_class_ids:
            return query.filter(Alert.id == -1)
        return query.filter(Student.class_id.in_(managed_class_ids))

    if current_user.role == UserRole.STUDENT:
        if not current_user.student_id:
            return query.filter(Alert.id == -1)
        return query.filter(Alert.student_id == current_user.student_id)

    return query.filter(Alert.id == -1)


def _get_alert_or_404(db: Session, alert_id: int) -> Alert:
    alert = (
        db.query(Alert)
        .options(
            joinedload(Alert.student).joinedload(Student.class_info),
            joinedload(Alert.rule),
            joinedload(Alert.records).joinedload(AlertRecord.handler).joinedload(User.student),
        )
        .filter(Alert.id == alert_id)
        .first()
    )
    if not alert:
        raise HTTPException(
            status_code=http_status.HTTP_404_NOT_FOUND,
            detail="预警不存在",
        )
    return alert


def _ensure_class_access(current_user: User, class_id: int) -> None:
    if current_user.role == UserRole.COUNSELOR and class_id not in current_user.get_managed_class_ids():
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="您没有权限访问该班级的数据",
        )


def _serialize_alert_list_item(alert: Alert) -> dict:
    return {
        "id": alert.id,
        "student_id": alert.student_id,
        "student_name": alert.student.name if alert.student else None,
        "student_no": alert.student.student_no if alert.student else None,
        "class_name": alert.student.class_info.name if alert.student and alert.student.class_info else None,
        "rule_id": alert.rule_id,
        "rule_name": alert.rule.name if alert.rule else None,
        "rule_type": _get_rule_display_type(alert.rule),
        "level": alert.level,
        "message": alert.message,
        "status": alert.status,
        "student_feedback": alert.student_feedback,
        "feedback_time": alert.feedback_time,
        "created_at": alert.created_at,
    }


def _serialize_alert_detail(alert: Alert) -> dict:
    records = sorted(alert.records or [], key=lambda item: item.created_at or datetime.min)
    return {
        "id": alert.id,
        "student_id": alert.student_id,
        "level": alert.level,
        "message": alert.message,
        "status": alert.status,
        "semester": alert.semester,
        "student_feedback": alert.student_feedback,
        "feedback_time": alert.feedback_time,
        "created_at": alert.created_at,
        "updated_at": alert.updated_at,
        "student": {
            "id": alert.student.id,
            "student_no": alert.student.student_no,
            "name": alert.student.name,
            "class_name": alert.student.class_info.name if alert.student.class_info else None,
            "phone": alert.student.phone,
            "email": alert.student.email,
        } if alert.student else None,
        "rule": {
            "id": alert.rule.id,
            "name": alert.rule.name,
            "code": alert.rule.code,
            "type": _get_rule_display_type(alert.rule),
            "level": alert.rule.level.value,
            "description": alert.rule.description,
            "conditions": alert.rule.conditions,
            "message_template": alert.rule.message_template,
        } if alert.rule else None,
        "records": [
            {
                "id": record.id,
                "handler_name": (
                    record.handler.student.name
                    if record.handler and record.handler.student
                    else (record.handler.username if record.handler else None)
                ),
                "action": record.action,
                "result": record.result,
                "created_at": record.created_at,
            }
            for record in records
        ],
    }


@router.get("", response_model=dict, summary="获取预警列表")
def get_alerts(
    page: int = 1,
    page_size: int = 20,
    status: Optional[AlertStatus] = None,
    level: Optional[AlertLevel] = None,
    class_id: Optional[int] = None,
    student_name: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = _scoped_alert_query(db, current_user)

    if status is not None:
        query = query.filter(Alert.status == status)

    if level is not None:
        query = query.filter(Alert.level == level)

    if class_id is not None:
        _ensure_class_access(current_user, class_id)
        query = query.filter(Student.class_id == class_id)

    if student_name:
        query = query.filter(Student.name.like(f"%{student_name}%"))

    total = query.count()
    alerts = (
        query.order_by(Alert.created_at.desc(), Alert.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {
        "items": [_serialize_alert_list_item(alert) for alert in alerts],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/statistics", summary="获取预警统计")
def get_alert_statistics(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = _scoped_alert_query(db, current_user)

    total = query.count()
    return {
        "total": total,
        "by_status": {
            "pending": query.filter(Alert.status == AlertStatus.PENDING).count(),
            "processing": query.filter(Alert.status == AlertStatus.PROCESSING).count(),
            "resolved": query.filter(Alert.status == AlertStatus.RESOLVED).count(),
            "ignored": query.filter(Alert.status == AlertStatus.IGNORED).count(),
        },
        "by_level": {
            "warning": query.filter(Alert.level == AlertLevel.WARNING).count(),
            "serious": query.filter(Alert.level == AlertLevel.SERIOUS).count(),
            "urgent": query.filter(Alert.level == AlertLevel.URGENT).count(),
        },
    }


@router.get("/export", summary="导出预警报表")
def export_alerts(
    status: Optional[AlertStatus] = None,
    level: Optional[AlertLevel] = None,
    class_id: Optional[int] = None,
    student_name: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_counselor_or_admin),
):
    query = _scoped_alert_query(db, current_user)

    if status is not None:
        query = query.filter(Alert.status == status)
    if level is not None:
        query = query.filter(Alert.level == level)
    if class_id is not None:
        _ensure_class_access(current_user, class_id)
        query = query.filter(Student.class_id == class_id)
    if student_name:
        query = query.filter(Student.name.like(f"%{student_name}%"))

    alerts = query.order_by(Alert.created_at.desc(), Alert.id.desc()).all()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "预警报表"

    headers = ["学号", "姓名", "班级", "预警规则", "预警级别", "生成时间", "当前状态"]
    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    widths = [16, 12, 20, 20, 12, 22, 12]

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        worksheet.column_dimensions[cell.column_letter].width = widths[column_index - 1]

    for row_index, alert in enumerate(alerts, start=2):
        worksheet.cell(row=row_index, column=1, value=alert.student.student_no if alert.student else "")
        worksheet.cell(row=row_index, column=2, value=alert.student.name if alert.student else "")
        worksheet.cell(
            row=row_index,
            column=3,
            value=alert.student.class_info.name if alert.student and alert.student.class_info else "",
        )
        worksheet.cell(row=row_index, column=4, value=alert.rule.name if alert.rule else "")
        worksheet.cell(row=row_index, column=5, value=_get_level_label(alert.level))
        worksheet.cell(
            row=row_index,
            column=6,
            value=alert.created_at.strftime("%Y-%m-%d %H:%M:%S") if alert.created_at else "",
        )
        worksheet.cell(row=row_index, column=7, value=_get_status_label(alert.status))

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"学业预警报表_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/{alert_id}", response_model=dict, summary="获取预警详情")
def get_alert(
    alert_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    alert = _get_alert_or_404(db, alert_id)
    if not check_student_access(current_user, alert.student_id, db):
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="您没有权限查看该预警",
        )
    return _serialize_alert_detail(alert)


@router.post("/{alert_id}/feedback", summary="提交学生反馈")
def submit_alert_feedback(
    alert_id: int,
    data: AlertFeedbackRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_student_user),
):
    alert = _get_alert_or_404(db, alert_id)
    if alert.student_id != current_user.student_id:
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="您只能提交自己的预警反馈",
        )

    feedback = data.feedback.strip()
    if not feedback:
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail="反馈内容不能为空",
        )

    if alert.student_feedback and alert.student_feedback.strip():
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail="该预警已经提交过反馈，不能重复提交",
        )

    alert.student_feedback = feedback
    alert.feedback_time = datetime.now()
    db.commit()
    db.refresh(alert)

    return {
        "message": "反馈提交成功",
        "student_feedback": alert.student_feedback,
        "feedback_time": alert.feedback_time,
    }


@router.post("/{alert_id}/handle", summary="处理预警")
def handle_alert(
    alert_id: int,
    data: HandleRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_counselor_or_admin),
):
    alert = _get_alert_or_404(db, alert_id)
    if current_user.role == UserRole.COUNSELOR and not check_student_access(current_user, alert.student_id, db):
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="您没有权限处理该预警",
        )

    record = AlertRecord(
        alert_id=alert_id,
        handler_id=current_user.id,
        action=data.action.strip(),
        result=data.result.strip() if data.result else None,
    )
    db.add(record)

    if alert.status == AlertStatus.PENDING:
        alert.status = AlertStatus.PROCESSING

    db.commit()
    db.refresh(record)
    return {"message": "处理记录已保存", "record_id": record.id}


@router.put("/{alert_id}/status", summary="更新预警状态")
def update_alert_status(
    alert_id: int,
    status: AlertStatus,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_counselor_or_admin),
):
    alert = _get_alert_or_404(db, alert_id)
    if current_user.role == UserRole.COUNSELOR and not check_student_access(current_user, alert.student_id, db):
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="您没有权限更新该预警状态",
        )

    alert.status = status
    db.commit()
    return {"message": "状态更新成功"}


@router.post("/{alert_id}/resolve", summary="解决预警")
def resolve_alert(
    alert_id: int,
    result: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_counselor_or_admin),
):
    alert = _get_alert_or_404(db, alert_id)
    if current_user.role == UserRole.COUNSELOR and not check_student_access(current_user, alert.student_id, db):
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="您没有权限解决该预警",
        )

    db.add(
        AlertRecord(
            alert_id=alert_id,
            handler_id=current_user.id,
            action="resolve",
            result=result or "预警已解决",
        )
    )
    alert.status = AlertStatus.RESOLVED
    db.commit()
    return {"message": "预警已标记为已解决"}


@router.post("/{alert_id}/ignore", summary="忽略预警")
def ignore_alert(
    alert_id: int,
    reason: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_counselor_or_admin),
):
    alert = _get_alert_or_404(db, alert_id)
    if current_user.role == UserRole.COUNSELOR and not check_student_access(current_user, alert.student_id, db):
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="您没有权限忽略该预警",
        )

    db.add(
        AlertRecord(
            alert_id=alert_id,
            handler_id=current_user.id,
            action="ignore",
            result=reason or "预警已忽略",
        )
    )
    alert.status = AlertStatus.IGNORED
    db.commit()
    return {"message": "预警已忽略"}


@router.get("/by-class/{class_id}", summary="获取指定班级的预警")
def get_alerts_by_class(
    class_id: int,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == UserRole.STUDENT:
        return {"items": [], "total": 0, "page": page, "page_size": page_size}

    _ensure_class_access(current_user, class_id)
    query = _scoped_alert_query(db, current_user).filter(Student.class_id == class_id)
    total = query.count()
    alerts = (
        query.order_by(Alert.created_at.desc(), Alert.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return {
        "items": [_serialize_alert_list_item(alert) for alert in alerts],
        "total": total,
        "page": page,
        "page_size": page_size,
    }
